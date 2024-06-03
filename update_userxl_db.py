import openpyxl as xl

#to store the excel database
user_db = {}

def import_user_db(path):

    #open excel file and activate the user_db sheet
    wb = xl.load_workbook(path)
    sheet = wb.active

    #Max row extraction
    #max_col = sheet.max_column
    max_row = sheet.max_row

    #save the user_db from execl to user_db dict
    for i in range(2, max_row+1):
        user_db[str(sheet.cell(i,1).value)] = [str(sheet.cell(i,2).value),str(sheet.cell(i,3).value)]

    wb.save(path)
    wb.close()

    return user_db

#update existing User.  

#Save new user in user_db to excel user db
def save_new_user(path,user_db):

    #open excel file and activate the user_db sheet
    wb = xl.load_workbook(path)
    sheet = wb.active

    #Max row extraction
    #max_col = sheet.max_column
    #max_row = sheet.max_row
    max_row_new = len(user_db)

    #save the user_db from execl to user_db dict
    sheet.cell(max_row_new+1,1).value = list(user_db.keys())[max_row_new-1]
    sheet.cell(max_row_new+1,2).value = user_db[str(sheet.cell(max_row_new+1,1).value)][0]
    sheet.cell(max_row_new+1,3).value = user_db[str(sheet.cell(max_row_new+1,1).value)][1] 


    wb.save(path)
    wb.close()